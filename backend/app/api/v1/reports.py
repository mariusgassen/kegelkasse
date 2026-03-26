"""
Report export — admin-only endpoint to generate an Excel (.xlsx) or PDF workbook
covering member accounts, penalties, transactions, and evening summaries.
Supports optional year filtering (omit for all-time).
"""
import io
import logging
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from api.deps import require_club_admin
from core.database import get_db
from models.club import Club
from models.evening import Evening, EveningPlayer, RegularMember
from models.game import Game
from models.payment import ClubExpense, MemberPayment
from models.penalty import PenaltyLog
from models.user import User

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/reports", tags=["reports"])

_MONTH_ABB_DE = ["Jan", "Feb", "Mrz", "Apr", "Mai", "Jun", "Jul", "Aug", "Sep", "Okt", "Nov", "Dez"]

# ── Styling helpers ──────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1E293B")
_ALT_FILL = PatternFill("solid", fgColor="0F172A")
_HEADER_FONT = Font(bold=True, color="E2E8F0")
_TOTAL_FONT = Font(bold=True, color="F1F5F9")
_TOTAL_FILL = PatternFill("solid", fgColor="334155")


def _write_header(ws, cols: list[str]) -> None:
    """Write a styled header row to the worksheet."""
    for ci, title in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20


def _auto_width(ws, min_w: int = 10, max_w: int = 50) -> None:
    """Auto-size all columns based on their content."""
    for col_cells in ws.columns:
        length = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, min_w), max_w)


def _fmt_date(dt: Optional[datetime]) -> str:
    if dt is None:
        return ""
    return dt.strftime("%d.%m.%Y")


def _fmt_euro(v: float) -> str:
    return f"{v:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")


def _penalty_euro(log: PenaltyLog) -> float:
    if log.mode == "euro":
        return log.amount
    if log.unit_amount is not None:
        return log.amount * log.unit_amount
    return 0.0


# ── Main export endpoint ──────────────────────────────────────────────────────

@router.get("/export")
def export_report(
    year: Optional[int] = Query(None, description="Filter to a specific year; omit for all-time"),
    fmt: str = Query("xlsx", alias="format", description="Output format: xlsx or pdf"),
    db: Session = Depends(get_db),
    user: User = Depends(require_club_admin),
):
    """Download a full report for the club (admin only). Supports xlsx and pdf formats."""
    if fmt not in ("xlsx", "pdf"):
        raise HTTPException(400, "format must be 'xlsx' or 'pdf'")
    club = db.query(Club).filter(Club.id == user.club_id).first()
    club_name = club.name if club else "Kegelclub"

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # ── Load base data ────────────────────────────────────────────────────────

    members = (
        db.query(RegularMember)
        .filter(RegularMember.club_id == user.club_id, RegularMember.is_active == True)
        .order_by(RegularMember.is_guest, RegularMember.name)
        .all()
    )
    member_by_id: dict[int, RegularMember] = {m.id: m for m in members}

    # Evenings (optionally year-filtered)
    evening_q = db.query(Evening).filter(Evening.club_id == user.club_id)
    if year:
        evening_q = evening_q.filter(
            Evening.date >= datetime(year, 1, 1, tzinfo=timezone.utc),
            Evening.date < datetime(year + 1, 1, 1, tzinfo=timezone.utc),
        )
    evenings = evening_q.order_by(Evening.date).all()
    evening_ids = [e.id for e in evenings]
    evening_by_id: dict[int, Evening] = {e.id: e for e in evenings}

    # EveningPlayers for these evenings
    players_in_scope = (
        db.query(EveningPlayer)
        .filter(EveningPlayer.evening_id.in_(evening_ids))
        .all()
    ) if evening_ids else []
    player_by_id: dict[int, EveningPlayer] = {p.id: p for p in players_in_scope}

    # PenaltyLog for these evenings (non-deleted)
    penalty_logs = (
        db.query(PenaltyLog)
        .filter(PenaltyLog.evening_id.in_(evening_ids), PenaltyLog.is_deleted == False)
        .order_by(PenaltyLog.evening_id, PenaltyLog.id)
        .all()
    ) if evening_ids else []

    # Payments (optionally year-filtered)
    payment_q = db.query(MemberPayment).filter(MemberPayment.club_id == user.club_id)
    if year:
        payment_q = payment_q.filter(
            MemberPayment.created_at >= datetime(year, 1, 1, tzinfo=timezone.utc),
            MemberPayment.created_at < datetime(year + 1, 1, 1, tzinfo=timezone.utc),
        )
    payments = payment_q.order_by(MemberPayment.created_at).all()

    # Expenses (optionally year-filtered)
    expense_q = db.query(ClubExpense).filter(ClubExpense.club_id == user.club_id)
    if year:
        expense_q = expense_q.filter(
            ClubExpense.created_at >= datetime(year, 1, 1, tzinfo=timezone.utc),
            ClubExpense.created_at < datetime(year + 1, 1, 1, tzinfo=timezone.utc),
        )
    expenses = expense_q.order_by(ClubExpense.created_at).all()

    # Games for these evenings
    games = (
        db.query(Game)
        .filter(Game.evening_id.in_(evening_ids), Game.is_deleted == False)
        .all()
    ) if evening_ids else []

    # ── Build aggregated lookups ──────────────────────────────────────────────

    # penalty euro per player_id and per regular_member_id (absences)
    penalty_by_player: dict[int, float] = {}
    penalty_by_member_absence: dict[int, float] = {}
    for log in penalty_logs:
        euro = _penalty_euro(log)
        if log.player_id is not None:
            penalty_by_player[log.player_id] = penalty_by_player.get(log.player_id, 0.0) + euro
        elif log.regular_member_id is not None:
            penalty_by_member_absence[log.regular_member_id] = (
                penalty_by_member_absence.get(log.regular_member_id, 0.0) + euro
            )

    # map regular_member_id → list of player_ids (in scope)
    member_player_ids: dict[int, list[int]] = {}
    for p in players_in_scope:
        if p.regular_member_id is not None:
            member_player_ids.setdefault(p.regular_member_id, []).append(p.id)

    def member_penalty_total(mid: int) -> float:
        total = sum(penalty_by_player.get(pid, 0.0) for pid in member_player_ids.get(mid, []))
        total += penalty_by_member_absence.get(mid, 0.0)
        return round(total, 2)

    payments_by_member: dict[int, float] = {}
    for p in payments:
        payments_by_member[p.regular_member_id] = payments_by_member.get(p.regular_member_id, 0.0) + p.amount

    # king per evening: player that has is_king=True
    king_per_evening: dict[int, str] = {}
    for p in players_in_scope:
        if p.is_king:
            king_per_evening[p.evening_id] = p.name

    # games per evening count
    games_per_evening: dict[int, int] = {}
    for g in games:
        games_per_evening[g.evening_id] = games_per_evening.get(g.evening_id, 0) + 1

    # penalty total per evening
    penalty_per_evening: dict[int, float] = {}
    for log in penalty_logs:
        eid = log.evening_id
        penalty_per_evening[eid] = penalty_per_evening.get(eid, 0.0) + _penalty_euro(log)

    # players per evening count
    players_per_evening: dict[int, int] = {}
    for p in players_in_scope:
        players_per_evening[p.evening_id] = players_per_evening.get(p.evening_id, 0) + 1

    # ── Shared derived data (used by both Excel and PDF) ──────────────────────

    period_label = str(year) if year else "Gesamt"
    total_penalties = round(sum(_penalty_euro(l) for l in penalty_logs), 2)
    total_payments = round(sum(p.amount for p in payments), 2)
    total_expenses = round(sum(e.amount for e in expenses), 2)
    kassenstand = round(total_payments - total_penalties - total_expenses, 2)
    regular_members = [m for m in members if not m.is_guest]

    # Bookings: payments + expenses merged and sorted by date
    booking_rows: list[tuple[Optional[datetime], str, str, float, str]] = []
    for p in payments:
        m = member_by_id.get(p.regular_member_id)
        name = (m.nickname or m.name) if m else f"#{p.regular_member_id}"
        booking_rows.append((p.created_at, "Einzahlung", name, p.amount, p.note or ""))
    for e in expenses:
        dt = e.date or e.created_at
        booking_rows.append((dt, "Ausgabe", e.description, -e.amount, ""))
    booking_rows.sort(key=lambda r: r[0] or datetime.min)

    # per (member_id, penalty_type_name) → (count, total_euro)
    pen_by_member_type: dict[tuple[int, str], tuple[int, float]] = {}
    for log in penalty_logs:
        if log.player_id is not None:
            ep = player_by_id.get(log.player_id)
            mid = ep.regular_member_id if ep else None
        else:
            mid = log.regular_member_id
        if mid is None:
            continue
        key = (mid, log.penalty_type_name or "Unbekannt")
        cnt, tot = pen_by_member_type.get(key, (0, 0.0))
        pen_by_member_type[key] = (cnt + 1, round(tot + _penalty_euro(log), 2))

    # Monthly penalties per member: {member_id: {(year, month): total_euro}}
    monthly_pen_by_member: dict[int, dict[tuple[int, int], float]] = defaultdict(lambda: defaultdict(float))
    for log in penalty_logs:
        ev = evening_by_id.get(log.evening_id)
        if ev is None:
            continue
        ym = (ev.date.year, ev.date.month)
        euro = _penalty_euro(log)
        if log.player_id is not None:
            ep = player_by_id.get(log.player_id)
            mid = ep.regular_member_id if ep else None
        else:
            mid = log.regular_member_id
        if mid is not None:
            monthly_pen_by_member[mid][ym] += euro

    all_months_set: set[tuple[int, int]] = set()
    for by_month in monthly_pen_by_member.values():
        all_months_set.update(by_month.keys())
    sorted_months = sorted(all_months_set)
    single_year = len({ym[0] for ym in sorted_months}) <= 1

    def _month_label(ym: tuple[int, int]) -> str:
        label = _MONTH_ABB_DE[ym[1] - 1]
        return label if single_year else f"{label} '{ym[0] % 100:02d}"

    # Guests with any penalty activity in scope
    guest_members = [m for m in members if m.is_guest and m.id in monthly_pen_by_member]

    year_suffix = f"_{year}" if year else ""

    if fmt == "pdf":
        # ── PDF output ────────────────────────────────────────────────────────
        buf = _build_pdf(
            club_name=club_name,
            period_label=period_label,
            total_penalties=total_penalties,
            total_payments=total_payments,
            total_expenses=total_expenses,
            kassenstand=kassenstand,
            evenings=evenings,
            regular_members=regular_members,
            member_penalty_total=member_penalty_total,
            payments_by_member=payments_by_member,
            booking_rows=booking_rows,
            pen_by_member_type=pen_by_member_type,
            penalty_per_evening=penalty_per_evening,
            players_per_evening=players_per_evening,
            games_per_evening=games_per_evening,
            king_per_evening=king_per_evening,
            monthly_pen_by_member=monthly_pen_by_member,
            sorted_months=sorted_months,
            guest_members=guest_members,
            month_label_fn=_month_label,
        )
        filename = f"kegelkasse_report{year_suffix}.pdf"
        logger.info("PDF report generated: club=%d user=%d year=%s", user.club_id, user.id, year)
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # ── Excel output ──────────────────────────────────────────────────────────

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # ── Sheet 1: Übersicht ────────────────────────────────────────────────────

    ws1 = wb.create_sheet("Übersicht")
    ws1.freeze_panes = "A2"

    ws1.cell(1, 1, "Vereinskasse — Bericht").font = Font(bold=True, size=14)
    ws1.cell(2, 1, "Verein").font = Font(bold=True)
    ws1.cell(2, 2, club_name)
    ws1.cell(3, 1, "Zeitraum").font = Font(bold=True)
    ws1.cell(3, 2, period_label)
    ws1.cell(4, 1, "Erstellt am").font = Font(bold=True)
    ws1.cell(4, 2, datetime.now().strftime("%d.%m.%Y %H:%M"))
    ws1.cell(5, 1, "")

    stats = [
        ("Strafen (gesamt)", _fmt_euro(total_penalties)),
        ("Einzahlungen (gesamt)", _fmt_euro(total_payments)),
        ("Ausgaben (gesamt)", _fmt_euro(total_expenses)),
        ("Kassenstand", _fmt_euro(kassenstand)),
        ("", ""),
        ("Abende", str(len(evenings))),
        ("Mitglieder (aktiv)", str(len(regular_members))),
    ]
    for ri, (label, value) in enumerate(stats, start=6):
        ws1.cell(ri, 1, label).font = Font(bold=True)
        ws1.cell(ri, 2, value)

    ws1.column_dimensions["A"].width = 26
    ws1.column_dimensions["B"].width = 20

    # ── Sheet 2: Mitglieder-Konten ────────────────────────────────────────────

    ws2 = wb.create_sheet("Mitglieder-Konten")
    ws2.freeze_panes = "A2"
    cols2 = ["Kegelname", "Name", "Strafen (€)", "Einzahlungen (€)", "Saldo (€)"]
    _write_header(ws2, cols2)

    for ri, m in enumerate(regular_members, start=2):
        display = m.nickname or m.name
        pen = member_penalty_total(m.id)
        pay = round(payments_by_member.get(m.id, 0.0), 2)
        bal = round(pay - pen, 2)
        row = [display, m.name, pen, pay, bal]
        for ci, val in enumerate(row, start=1):
            cell = ws2.cell(ri, ci, val)
            if ci in (3, 4, 5):
                cell.number_format = '#,##0.00 "€"'
                if ci == 5:
                    cell.font = Font(color="FF4444" if bal < 0 else "4ADE80")

    # Totals row
    tr = len(regular_members) + 2
    ws2.cell(tr, 1, "GESAMT").font = _TOTAL_FONT
    for ci in (3, 4, 5):
        ws2.cell(tr, ci).fill = _TOTAL_FILL
    for ci in range(1, 6):
        ws2.cell(tr, ci).font = _TOTAL_FONT
        ws2.cell(tr, ci).fill = _TOTAL_FILL
    tot_pen = round(sum(member_penalty_total(m.id) for m in regular_members), 2)
    tot_pay = round(sum(payments_by_member.get(m.id, 0.0) for m in regular_members), 2)
    ws2.cell(tr, 3, tot_pen).number_format = '#,##0.00 "€"'
    ws2.cell(tr, 4, tot_pay).number_format = '#,##0.00 "€"'
    ws2.cell(tr, 5, round(tot_pay - tot_pen, 2)).number_format = '#,##0.00 "€"'
    _auto_width(ws2)

    # ── Sheet 3: Buchungen ────────────────────────────────────────────────────

    ws3 = wb.create_sheet("Buchungen")
    ws3.freeze_panes = "A2"
    _write_header(ws3, ["Datum", "Typ", "Mitglied / Beschreibung", "Betrag (€)", "Notiz / Buchungstext"])

    for ri, (dt, typ, desc, amt, note) in enumerate(booking_rows, start=2):
        ws3.cell(ri, 1, _fmt_date(dt))
        ws3.cell(ri, 2, typ)
        ws3.cell(ri, 3, desc)
        cell_amt = ws3.cell(ri, 4, amt)
        cell_amt.number_format = '#,##0.00 "€"'
        if amt < 0:
            cell_amt.font = Font(color="FF4444")
        ws3.cell(ri, 5, note)

    _auto_width(ws3)

    # ── Sheet 4: Strafen nach Person ──────────────────────────────────────────

    ws4 = wb.create_sheet("Strafen nach Person")
    ws4.freeze_panes = "A2"
    _write_header(ws4, ["Kegelname", "Name", "Strafen-Typ", "Anzahl", "Betrag (€)"])

    ri = 2
    for m in regular_members:
        entries = [(k[1], v[0], v[1]) for k, v in pen_by_member_type.items() if k[0] == m.id]
        entries.sort(key=lambda x: x[0])
        for pname, cnt, total in entries:
            ws4.cell(ri, 1, m.nickname or m.name)
            ws4.cell(ri, 2, m.name)
            ws4.cell(ri, 3, pname)
            ws4.cell(ri, 4, cnt)
            ws4.cell(ri, 5, total).number_format = '#,##0.00 "€"'
            ri += 1

    _auto_width(ws4)

    # ── Sheet 5: Strafen nach Abend ───────────────────────────────────────────

    ws5 = wb.create_sheet("Strafen nach Abend")
    ws5.freeze_panes = "A2"
    _write_header(ws5, ["Datum", "Abend", "Mitglied", "Strafen-Typ", "Modus", "Anzahl", "Betrag (€)"])

    for log in penalty_logs:
        ev = evening_by_id.get(log.evening_id)
        if log.player_id is not None:
            ep = player_by_id.get(log.player_id)
            member_name = ep.name if ep else log.player_name or ""
        else:
            m = member_by_id.get(log.regular_member_id) if log.regular_member_id else None
            member_name = (m.nickname or m.name) if m else (log.player_name or "")
        ws5.cell(ws5.max_row + 1 if ws5.max_row > 1 else 2, 1)
        row_idx = ws5.max_row if ws5.max_row > 1 else 2
        ws5.cell(row_idx, 1, _fmt_date(ev.date) if ev else "")
        ws5.cell(row_idx, 2, ev.venue or "" if ev else "")
        ws5.cell(row_idx, 3, member_name)
        ws5.cell(row_idx, 4, log.penalty_type_name or "")
        ws5.cell(row_idx, 5, log.mode or "")
        ws5.cell(row_idx, 6, int(log.amount) if log.mode == "count" else "")
        ws5.cell(row_idx, 7, round(_penalty_euro(log), 2)).number_format = '#,##0.00 "€"'

    _auto_width(ws5)

    # ── Sheet 6: Abende ───────────────────────────────────────────────────────

    ws6 = wb.create_sheet("Abende")
    ws6.freeze_panes = "A2"
    _write_header(ws6, ["Datum", "Ort", "Status", "Spieler", "Spiele", "Strafen (€)", "König"])

    for ri, ev in enumerate(reversed(evenings), start=2):
        ws6.cell(ri, 1, _fmt_date(ev.date))
        ws6.cell(ri, 2, ev.venue or "")
        ws6.cell(ri, 3, "Archiviert" if ev.is_closed else "Offen")
        ws6.cell(ri, 4, players_per_evening.get(ev.id, 0))
        ws6.cell(ri, 5, games_per_evening.get(ev.id, 0))
        ws6.cell(ri, 6, round(penalty_per_evening.get(ev.id, 0.0), 2)).number_format = '#,##0.00 "€"'
        ws6.cell(ri, 7, king_per_evening.get(ev.id, ""))

    _auto_width(ws6)

    # ── Sheet 7: Monatsübersicht ──────────────────────────────────────────────

    ws7 = wb.create_sheet("Monatsübersicht")
    ws7.freeze_panes = "B2"
    month_cols = [_month_label(ym) for ym in sorted_months]
    _write_header(ws7, ["Spieler"] + month_cols + ["Strafen Ges.", "Einzahlungen", "Kontostand"])

    def _write_monthly_row(ws, ri: int, m, label_override: str = "") -> None:
        display = label_override or (m.nickname or m.name)
        ws.cell(ri, 1, display)
        for ci, ym in enumerate(sorted_months, start=2):
            v = round(monthly_pen_by_member[m.id].get(ym, 0.0), 2)
            cell = ws.cell(ri, ci, v if v else "")
            if v:
                cell.number_format = '#,##0.00 "€"'
        base = len(sorted_months)
        pen = member_penalty_total(m.id)
        pay = round(payments_by_member.get(m.id, 0.0), 2)
        ws.cell(ri, base + 2, pen).number_format = '#,##0.00 "€"'
        ws.cell(ri, base + 3, pay).number_format = '#,##0.00 "€"'
        ws.cell(ri, base + 4, round(pay - pen, 2)).number_format = '#,##0.00 "€"'

    ri7 = 2
    # Stammspieler section header
    if regular_members:
        h_cell = ws7.cell(ri7, 1, "— Stammspieler —")
        h_cell.font = Font(bold=True, color="E8A020")
        ri7 += 1
        for m in regular_members:
            _write_monthly_row(ws7, ri7, m)
            ri7 += 1

    # Guests section header
    if guest_members:
        ws7.cell(ri7, 1, "")
        ri7 += 1
        h_cell = ws7.cell(ri7, 1, "— Gäste —")
        h_cell.font = Font(bold=True, color="E8A020")
        ri7 += 1
        for m in guest_members:
            _write_monthly_row(ws7, ri7, m)
            ri7 += 1

    _auto_width(ws7)

    # ── Stream Excel ──────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"kegelkasse_report{year_suffix}.xlsx"
    logger.info("Excel report generated: club=%d user=%d year=%s", user.club_id, user.id, year)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── PDF builder ───────────────────────────────────────────────────────────────

def _build_pdf(
    *,
    club_name: str,
    period_label: str,
    total_penalties: float,
    total_payments: float,
    total_expenses: float,
    kassenstand: float,
    evenings: list,
    regular_members: list,
    member_penalty_total,
    payments_by_member: dict,
    booking_rows: list,
    pen_by_member_type: dict,
    penalty_per_evening: dict,
    players_per_evening: dict,
    games_per_evening: dict,
    king_per_evening: dict,
    monthly_pen_by_member: dict,
    sorted_months: list,
    guest_members: list,
    month_label_fn,
) -> io.BytesIO:
    # Load Liberation Sans — metrically identical to Helvetica, supports € and full Unicode
    _font_dir = Path(__file__).parent.parent.parent / "fonts"
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_font("Liberation", "", str(_font_dir / "LiberationSans-Regular.ttf"))
    pdf.add_font("Liberation", "B", str(_font_dir / "LiberationSans-Bold.ttf"))
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_margins(15, 15, 15)

    def _s(txt: str) -> str:
        """Pass-through; Liberation Sans is a Unicode font — no sanitisation needed."""
        return str(txt)

    def _h1(txt: str) -> None:
        pdf.set_font("Liberation", "B", 14)
        pdf.set_text_color(232, 160, 32)  # amber
        pdf.cell(0, 8, _s(txt), new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)
        pdf.ln(2)

    def _h2(txt: str) -> None:
        pdf.set_font("Liberation", "B", 10)
        pdf.set_fill_color(30, 41, 59)
        pdf.set_text_color(226, 232, 240)
        pdf.cell(0, 7, _s(f"  {txt}"), new_x="LMARGIN", new_y="NEXT", fill=True)
        pdf.set_text_color(0, 0, 0)
        pdf.ln(1)

    def _row(cols: list[tuple[str, float]], fill: bool = False) -> None:
        pdf.set_font("Liberation", size=8)
        if fill:
            pdf.set_fill_color(240, 240, 240)
        for txt, w in cols:
            pdf.cell(w, 6, _s(str(txt)), border=0, fill=fill)
        pdf.ln()

    def _header_row(cols: list[tuple[str, float]]) -> None:
        pdf.set_font("Liberation", "B", 8)
        pdf.set_fill_color(30, 41, 59)
        pdf.set_text_color(226, 232, 240)
        for txt, w in cols:
            pdf.cell(w, 7, _s(txt), border=0, fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)

    # ── Cover / Summary ──────────────────────────────────────────────────────
    pdf.add_page()
    _h1(f"Kassenbericht - {club_name}")
    pdf.set_font("Liberation", size=9)
    pdf.cell(50, 6, "Zeitraum:")
    pdf.cell(0, 6, _s(period_label), new_x="LMARGIN", new_y="NEXT")
    pdf.cell(50, 6, "Erstellt am:")
    pdf.cell(0, 6, datetime.now().strftime("%d.%m.%Y %H:%M"), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    summary = [
        ("Strafen (gesamt)", _fmt_euro(total_penalties)),
        ("Einzahlungen (gesamt)", _fmt_euro(total_payments)),
        ("Ausgaben (gesamt)", _fmt_euro(total_expenses)),
        ("Kassenstand", _fmt_euro(kassenstand)),
        ("Abende", str(len(evenings))),
        ("Mitglieder (aktiv)", str(len(regular_members))),
    ]
    for label, value in summary:
        pdf.set_font("Liberation", "B", 9)
        pdf.cell(65, 6, _s(label + ":"))
        pdf.set_font("Liberation", size=9)
        pdf.cell(0, 6, _s(value), new_x="LMARGIN", new_y="NEXT")

    # ── Mitglieder-Konten ────────────────────────────────────────────────────
    pdf.ln(4)
    _h2("Mitglieder-Konten")
    _header_row([("Kegelname", 45), ("Name", 50), ("Strafen", 30), ("Einz.", 30), ("Saldo", 25)])
    for i, m in enumerate(regular_members):
        pen = member_penalty_total(m.id)
        pay = round(payments_by_member.get(m.id, 0.0), 2)
        bal = round(pay - pen, 2)
        _row([
            (m.nickname or m.name, 45),
            (m.name, 50),
            (_fmt_euro(pen), 30),
            (_fmt_euro(pay), 30),
            (_fmt_euro(bal), 25),
        ], fill=(i % 2 == 1))

    # Totals
    tot_pen = round(sum(member_penalty_total(m.id) for m in regular_members), 2)
    tot_pay = round(sum(payments_by_member.get(m.id, 0.0) for m in regular_members), 2)
    pdf.set_font("Liberation", "B", 8)
    pdf.set_fill_color(30, 41, 59)
    pdf.set_text_color(241, 245, 249)
    for txt, w in [("GESAMT", 95), (_fmt_euro(tot_pen), 30), (_fmt_euro(tot_pay), 30), (_fmt_euro(tot_pay - tot_pen), 25)]:
        pdf.cell(w, 7, _s(txt), fill=True)
    pdf.ln()
    pdf.set_text_color(0, 0, 0)

    # ── Buchungen ────────────────────────────────────────────────────────────
    pdf.add_page()
    _h2("Buchungen")
    _header_row([("Datum", 22), ("Typ", 25), ("Mitglied / Beschreibung", 80), ("Betrag", 28), ("Notiz", 25)])
    for i, (dt, typ, desc, amt, note) in enumerate(booking_rows):
        _row([
            (_fmt_date(dt), 22),
            (typ, 25),
            (desc[:35], 80),
            (_fmt_euro(abs(amt)) + (" +" if amt >= 0 else " -"), 28),
            (note[:20], 25),
        ], fill=(i % 2 == 1))

    # ── Strafen nach Person ──────────────────────────────────────────────────
    pdf.add_page()
    _h2("Strafen nach Person")
    _header_row([("Kegelname", 40), ("Strafen-Typ", 70), ("Anz.", 15), ("Betrag", 25)])
    ri = 0
    for m in regular_members:
        entries = [(k[1], v[0], v[1]) for k, v in pen_by_member_type.items() if k[0] == m.id]
        entries.sort(key=lambda x: x[0])
        for pname, cnt, total in entries:
            _row([(m.nickname or m.name, 40), (pname, 70), (str(cnt), 15), (_fmt_euro(total), 25)], fill=(ri % 2 == 1))
            ri += 1

    # ── Abende ───────────────────────────────────────────────────────────────
    pdf.add_page()
    _h2("Abende")
    _header_row([("Datum", 22), ("Ort", 50), ("Status", 25), ("Spieler", 18), ("Spiele", 18), ("Strafen", 25), ("König", 22)])
    for i, ev in enumerate(reversed(evenings)):
        _row([
            (_fmt_date(ev.date), 22),
            ((ev.venue or "")[:22], 50),
            ("Arch." if ev.is_closed else "Offen", 25),
            (str(players_per_evening.get(ev.id, 0)), 18),
            (str(games_per_evening.get(ev.id, 0)), 18),
            (_fmt_euro(penalty_per_evening.get(ev.id, 0.0)), 25),
            ((king_per_evening.get(ev.id, ""))[:12], 22),
        ], fill=(i % 2 == 1))

    # ── Monatsübersicht ───────────────────────────────────────────────────────
    if sorted_months:
        pdf.add_page(orientation="L")  # landscape for wide table
        _h2("Monatsübersicht")

        # Calculate column widths to fit landscape page (267mm usable)
        name_w = 45.0
        fixed_w = 25.0 + 25.0 + 25.0  # Strafen Ges. + Einzahlungen + Kontostand
        available_w = 267.0 - name_w - fixed_w
        n_months = len(sorted_months)
        mon_w = min(22.0, round(available_w / n_months, 1)) if n_months else 20.0

        # Header
        h_cols = [("Spieler", name_w)] + [(month_label_fn(ym), mon_w) for ym in sorted_months]
        h_cols += [("Strafen Ges.", 25), ("Einzahlungen", 25), ("Kontostand", 25)]
        _header_row(h_cols)

        def _monthly_member_row(m, fill: bool) -> None:
            display = m.nickname or m.name
            pen = member_penalty_total(m.id)
            pay = round(payments_by_member.get(m.id, 0.0), 2)
            cols = [(display[:20], name_w)]
            for ym in sorted_months:
                v = monthly_pen_by_member[m.id].get(ym, 0.0)
                cols.append((_fmt_euro(v) if v else "", mon_w))
            cols += [(_fmt_euro(pen), 25), (_fmt_euro(pay), 25), (_fmt_euro(pay - pen), 25)]
            _row(cols, fill=fill)

        # Stammspieler
        if regular_members:
            pdf.set_font("Liberation", "B", 8)
            pdf.set_text_color(232, 160, 32)
            pdf.cell(0, 6, _s("Stammspieler"), new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(0, 0, 0)
            for i, m in enumerate(regular_members):
                _monthly_member_row(m, fill=(i % 2 == 1))

        # Guests
        if guest_members:
            pdf.ln(2)
            pdf.set_font("Liberation", "B", 8)
            pdf.set_text_color(232, 160, 32)
            pdf.cell(0, 6, _s("Gäste"), new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(0, 0, 0)
            for i, m in enumerate(guest_members):
                _monthly_member_row(m, fill=(i % 2 == 1))

    buf = io.BytesIO(pdf.output())
    buf.seek(0)
    return buf
